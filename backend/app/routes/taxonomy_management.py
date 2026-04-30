import csv
import hashlib
import io
import uuid
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from psycopg.rows import dict_row
from pydantic import BaseModel

from app.config.constants import DEFAULT_TENANT_ID
from app.infra.db import get_tenant_cursor
from app.middleware.permissions import require_minimum_role

router = APIRouter(prefix="/v1/taxonomy", tags=["taxonomy"])

analyst_gate = Depends(require_minimum_role("analyst"))
lead_gate = Depends(require_minimum_role("delivery_lead"))


# ── Pydantic bodies ───────────────────────────────────────────

class TaxonomyCreate(BaseModel):
    name: str
    description: Optional[str] = None
    source_type: str = "both"
    version: int = 1


class TaxonomyPatch(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    active: Optional[bool] = None


class PatchItem(BaseModel):
    risk_id: Optional[str] = None
    control_id: Optional[str] = None
    name: Optional[str] = None
    description: Optional[str] = None
    category: Optional[str] = None
    source: Optional[str] = None
    control_type: Optional[str] = None
    is_key: Optional[bool] = None


class ItemsPatch(BaseModel):
    item_type: str  # "risk" | "control"
    items: list[PatchItem]


# ── Helpers ───────────────────────────────────────────────────

def _parse_excel(content: bytes) -> tuple[list[dict], list[dict]]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    def sheet_rows(ws) -> list[dict]:
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row)})
        return rows

    risk_sheet = ctrl_sheet = None
    for name in wb.sheetnames:
        lower = name.lower()
        if "risk" in lower and risk_sheet is None:
            risk_sheet = wb[name]
        elif "control" in lower and ctrl_sheet is None:
            ctrl_sheet = wb[name]

    if risk_sheet is None and len(wb.sheetnames) >= 1:
        risk_sheet = wb[wb.sheetnames[0]]
    if ctrl_sheet is None and len(wb.sheetnames) >= 2:
        ctrl_sheet = wb[wb.sheetnames[1]]

    risks    = _normalise_risks(sheet_rows(risk_sheet))    if risk_sheet else []
    controls = _normalise_controls(sheet_rows(ctrl_sheet)) if ctrl_sheet else []
    return risks, controls


def _parse_csv(content: bytes) -> tuple[list[dict], list[dict]]:
    text = content.decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], []
    headers_lower = [h.lower() for h in rows[0].keys()]
    if any("risk" in h for h in headers_lower):
        return _normalise_risks(rows), []
    return [], _normalise_controls(rows)


def _normalise_risks(rows: list[dict]) -> list[dict]:
    out = []
    for r in rows:
        name = r.get("name") or r.get("Name") or r.get("Risk Name") or ""
        if not name.strip():
            continue
        out.append({
            "risk_id":     r.get("risk_id") or r.get("Risk ID") or f"R-{uuid.uuid4().hex[:6].upper()}",
            "category":    r.get("category") or r.get("Category") or "",
            "name":        name.strip(),
            "description": (r.get("description") or r.get("Description") or "").strip(),
            "source":      (r.get("source") or r.get("Source") or "EXT").strip().upper(),
        })
    return out


def _normalise_controls(rows: list[dict]) -> list[dict]:
    out = []
    for r in rows:
        name = r.get("control_name") or r.get("Name") or r.get("Control Name") or ""
        if not name.strip():
            continue
        raw_key = str(r.get("is_key") or r.get("Key Control") or "").strip().upper()
        out.append({
            "control_id":   r.get("control_id") or r.get("Control ID") or f"C-{uuid.uuid4().hex[:6].upper()}",
            "control_name": name.strip(),
            "description":  (r.get("description") or r.get("Description") or "").strip(),
            "control_type": (r.get("control_type") or r.get("Type") or "").strip() or None,
            "is_key":       raw_key in ("TRUE", "YES", "Y", "1"),
        })
    return out


# ── Endpoints ─────────────────────────────────────────────────

@router.get("")
async def list_taxonomies(request: Request):
    user = request.state.user
    tenant_id = user.get("tenantId", DEFAULT_TENANT_ID)
    async with get_tenant_cursor(tenant_id, row_factory=dict_row) as cur:
        await cur.execute(
            """SELECT id, name, version, source_type, risk_count, control_count,
                      active, file_name, uploaded_at, created_at
               FROM app.taxonomy_schemas
               WHERE tenant_id = %s
               ORDER BY created_at DESC""",
            (tenant_id,),
        )
        return await cur.fetchall()


@router.post("", status_code=201, dependencies=[analyst_gate])
async def create_taxonomy(body: TaxonomyCreate, request: Request):
    user = request.state.user
    tenant_id = user.get("tenantId", DEFAULT_TENANT_ID)
    tax_id = str(uuid.uuid4())
    async with get_tenant_cursor(tenant_id) as cur:
        await cur.execute(
            """INSERT INTO app.taxonomy_schemas
               (id, tenant_id, name, version, source_type, schema)
               VALUES (%s, %s, %s, %s, %s, %s)""",
            (tax_id, tenant_id, body.name, body.version, body.source_type, {}),
        )
    return {"id": tax_id}


@router.get("/{taxonomy_id}")
async def get_taxonomy(taxonomy_id: str, request: Request):
    user = request.state.user
    tenant_id = user.get("tenantId", DEFAULT_TENANT_ID)
    async with get_tenant_cursor(tenant_id, row_factory=dict_row) as cur:
        await cur.execute(
            """SELECT id, name, version, source_type, schema,
                      risks_data, controls_data, risk_count, control_count,
                      active, file_name, uploaded_at, created_at
               FROM app.taxonomy_schemas
               WHERE id = %s AND tenant_id = %s""",
            (taxonomy_id, tenant_id),
        )
        row = await cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Taxonomy not found")
    return row


@router.patch("/{taxonomy_id}", dependencies=[lead_gate])
async def patch_taxonomy(taxonomy_id: str, body: TaxonomyPatch, request: Request):
    user = request.state.user
    tenant_id = user.get("tenantId", DEFAULT_TENANT_ID)
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    if not updates:
        return {"id": taxonomy_id}
    updates["updated_at"] = "NOW()"
    set_clauses = ", ".join(f"{k} = %s" for k in updates if k != "updated_at")
    set_clauses += ", updated_at = NOW()"
    values = [v for k, v in updates.items() if k != "updated_at"] + [taxonomy_id, tenant_id]
    async with get_tenant_cursor(tenant_id) as cur:
        await cur.execute(
            f"UPDATE app.taxonomy_schemas SET {set_clauses} WHERE id = %s AND tenant_id = %s",
            values,
        )
    return {"id": taxonomy_id}


@router.delete("/{taxonomy_id}", status_code=204, dependencies=[lead_gate])
async def delete_taxonomy(taxonomy_id: str, request: Request):
    user = request.state.user
    tenant_id = user.get("tenantId", DEFAULT_TENANT_ID)
    async with get_tenant_cursor(tenant_id) as cur:
        await cur.execute(
            "DELETE FROM app.taxonomy_schemas WHERE id = %s AND tenant_id = %s",
            (taxonomy_id, tenant_id),
        )


@router.post("/{taxonomy_id}/upload", dependencies=[analyst_gate])
async def upload_taxonomy_file(taxonomy_id: str, request: Request, file: UploadFile = File(...)):
    user = request.state.user
    tenant_id = user.get("tenantId", DEFAULT_TENANT_ID)

    content = await file.read()
    sha256 = hashlib.sha256(content).hexdigest()

    # Check for duplicate
    async with get_tenant_cursor(tenant_id, row_factory=dict_row) as cur:
        await cur.execute(
            "SELECT id FROM app.taxonomy_schemas WHERE tenant_id = %s AND file_sha256 = %s AND id != %s",
            (tenant_id, sha256, taxonomy_id),
        )
        if await cur.fetchone():
            raise HTTPException(status_code=409, detail="Duplicate file — this file has already been uploaded.")

    fname = file.filename or ""
    if fname.endswith(".xlsx") or fname.endswith(".xls"):
        risks, controls = _parse_excel(content)
    else:
        risks, controls = _parse_csv(content)

    async with get_tenant_cursor(tenant_id) as cur:
        await cur.execute(
            """UPDATE app.taxonomy_schemas
               SET risks_data = %s, controls_data = %s,
                   risk_count = %s, control_count = %s,
                   file_name = %s, file_sha256 = %s,
                   uploaded_at = NOW(), updated_at = NOW()
               WHERE id = %s AND tenant_id = %s""",
            (risks, controls, len(risks), len(controls), fname, sha256, taxonomy_id, tenant_id),
        )

    return {"risks": len(risks), "controls": len(controls)}


@router.patch("/{taxonomy_id}/items", dependencies=[analyst_gate])
async def patch_taxonomy_items(taxonomy_id: str, body: ItemsPatch, request: Request):
    user = request.state.user
    tenant_id = user.get("tenantId", DEFAULT_TENANT_ID)

    async with get_tenant_cursor(tenant_id, row_factory=dict_row) as cur:
        await cur.execute(
            "SELECT risks_data, controls_data FROM app.taxonomy_schemas WHERE id = %s AND tenant_id = %s",
            (taxonomy_id, tenant_id),
        )
        row = await cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Taxonomy not found")

    if body.item_type == "risk":
        items_list: list[dict] = list(row["risks_data"])
        id_field = "risk_id"
    else:
        items_list = list(row["controls_data"])
        id_field = "control_id"

    patch_map = {p.model_dump()[id_field]: p.model_dump(exclude_none=True) for p in body.items if p.model_dump().get(id_field)}

    for i, item in enumerate(items_list):
        key = item.get(id_field)
        if key in patch_map:
            items_list[i] = {**item, **{k: v for k, v in patch_map[key].items() if v is not None}}

    if body.item_type == "risk":
        col, count_col = "risks_data", "risk_count"
    else:
        col, count_col = "controls_data", "control_count"

    async with get_tenant_cursor(tenant_id) as cur:
        await cur.execute(
            f"UPDATE app.taxonomy_schemas SET {col} = %s, {count_col} = %s, updated_at = NOW() WHERE id = %s AND tenant_id = %s",
            (items_list, len(items_list), taxonomy_id, tenant_id),
        )

    return {"updated": len(patch_map)}


@router.get("/{taxonomy_id}/export")
async def export_taxonomy(taxonomy_id: str, request: Request):
    user = request.state.user
    tenant_id = user.get("tenantId", DEFAULT_TENANT_ID)

    async with get_tenant_cursor(tenant_id, row_factory=dict_row) as cur:
        await cur.execute(
            "SELECT name, risks_data, controls_data FROM app.taxonomy_schemas WHERE id = %s AND tenant_id = %s",
            (taxonomy_id, tenant_id),
        )
        row = await cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Taxonomy not found")

    risks    = row["risks_data"]    or []
    controls = row["controls_data"] or []

    def generate():
        buf = io.StringIO()
        buf.write("=== RISKS ===\n")
        yield buf.getvalue()

        risk_fields = ["risk_id", "category", "name", "description", "source"]
        buf = io.StringIO()
        w = csv.DictWriter(buf, fieldnames=risk_fields, extrasaction="ignore")
        w.writeheader()
        yield buf.getvalue()
        for r in risks:
            buf = io.StringIO()
            w = csv.DictWriter(buf, fieldnames=risk_fields, extrasaction="ignore")
            w.writerow(r)
            yield buf.getvalue()

        buf = io.StringIO()
        buf.write("\n=== CONTROLS ===\n")
        yield buf.getvalue()

        ctrl_fields = ["control_id", "control_name", "description", "control_type", "is_key"]
        buf = io.StringIO()
        w = csv.DictWriter(buf, fieldnames=ctrl_fields, extrasaction="ignore")
        w.writeheader()
        yield buf.getvalue()
        for c in controls:
            buf = io.StringIO()
            w = csv.DictWriter(buf, fieldnames=ctrl_fields, extrasaction="ignore")
            w.writerow(c)
            yield buf.getvalue()

    safe_name = row["name"].replace(" ", "_")
    return StreamingResponse(
        generate(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_export.csv"'},
    )
